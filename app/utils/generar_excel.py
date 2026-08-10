import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import flet as ft        

class GenerarExcel:
    def __init__(self, page: ft.Page):
        self.page = page
        self.base = Path(__file__).resolve().parent # obtener la ruta del directorio actual del archivo Python
        self.ruta_log = self.base / "logs" / "app.log" # definir la ruta del archivo de log

    def generar_excel(self):
        wb = openpyxl.Workbook() # crear un nuevo libro de trabajo
        ws = wb.active # seleccionar la hoja activa

        cabecera = ["Fecha y Hora", "Nivel de log", "Mensaje"] # definir la cabecera
        ws.append(cabecera) # agregar la cabecera al archivo

        with open(self.ruta_log, encoding="utf-8") as archivo_log: # abrir el archivo de log en modo lectura
            for linea in archivo_log: # iterar sobre cada línea del archivo de log
                partes = linea.strip().split(" | ") # separar la línea en partes usando " | " como delimitador
                if len(partes) == 3: # verificar que la línea tenga exactamente 3 partes
                    ws.append(partes) # agregar las partes como una nueva fila en el archivo Excel

        for celda in ws[1]: # iterar sobre las celdas de la primera fila (cabecera)
            celda.font = Font(bold=True, color="FFFFFF") # aplicar formato de fuente en negrita y color blanco
            celda.fill = PatternFill(fill_type="solid", fgColor="2E4057") # aplicar color de fondo azul oscuro
            celda.alignment = Alignment(horizontal="center", vertical="center") # alinear el texto al centro

        ws.column_dimensions["A"].width = 20 # ajustar el ancho de la columna A
        ws.column_dimensions["B"].width = 15 # ajustar el ancho de la columna B
        ws.column_dimensions["C"].width = 50 # ajustar el ancho de la columna C

        ruta_excel = self.base / "logging.xlsx"
        wb.save(ruta_excel) # guardar el archivo Excel
        alerta = ft.AlertDialog(
            title=ft.Text("Exito"),
            content=ft.Text(f"Archivo '{ruta_excel}' creado con éxito."),
            actions=[
                ft.TextButton("Cerrar", on_click=lambda e: self.page.pop_dialog())
            ]
        )
        self.page.show_dialog(alerta)
